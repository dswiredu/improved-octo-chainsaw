import logging

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

from ps.parser import ChartTableFormatter
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport
from claret.accountant_report.reports.tax_prep import TaxPrepReport
from claret.accountant_report.report_utils import (
    generate_group_totals,
    set_pagination_size,
    set_trx_payload_custom_date_range,
    update_cell_range_values,
)
from claret.accountant_report import lookup

LOG = logging.getLogger(__name__)


class CapitalTransactions:
    """Class to generate Capital Transactions report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.payload_filename = "capital_transactions_payload.json"
        self.sheet_name = "Capital Transactions"
        self.payload_file = report.get_payload_file(self.payload_filename)
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)
        self.report_header = (
            f"Reporting Currency: {self.account_info['reporting_currency_name']}"
        )

        self.trx_id = "Transaction ID"
        self.currency_field = "transaction_currency"
        self.ttype_field = "d1g1t Transaction Type"
        self.effect_on_bv_ioc = "effect_on_bv_ioc"
        self.effect_on_bv_rc = "effect_on_bv_rc"
        self.realized_gl_rc = "realized_gl_rc"
        self.realized_gl_ioc = "realized_gl_ioc"
        self.book_value_rc = "book_value_rc"
        self.book_value_ioc = "book_value_ioc"

        self.grp_fields = [
            self.currency_field,
            "Name2",
            "Name3",
        ]

        self.num_cols = [
            self.effect_on_bv_rc,
            self.effect_on_bv_ioc,
            self.realized_gl_rc,
            self.realized_gl_ioc,
            self.book_value_rc,
            self.book_value_ioc,
        ]

        self.sum_columns = [
            "Asset Account Local",
            "Cash Account Local",
            "Realized G/L Account Local",
            "Asset Account Rep. Cur.",
            "Cash Account Rep. Cur.",
            "Realized G/L Account Rep. Cur.",
        ]

        self.report_columns = [
            self.currency_field,
            "Name2",
            "Name3",
            "Open Date",
            "Close Date",
            "Quantity",
            "Description",
            "Asset Account Local",
            "Cash Account Local",
            "Realized G/L Account Local",
            "Cost Forex",
            "Asset Account Rep. Cur.",
            "Proceeds Forex",
            "Cash Account Rep. Cur.",
            "Realized G/L Account Rep. Cur.",
            "legend",
        ]

        # Using list for now cos client still needs to confirm.
        self.opening_transactions = [
            "Buy",
        ]

        self.closing_transactions = [
            "Sell",
        ]

        self.report_ttypes = self.opening_transactions + self.closing_transactions

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        set_pagination_size(payload)
        set_trx_payload_custom_date_range(
            payload, start=self.report.trxs_start_date, end=self.end_date
        )
        return payload

    def set_report_dates(self, df: pd.DataFrame) -> pd.Series:
        open_idx = df[self.ttype_field].isin(self.opening_transactions)
        dates = df["Date"].dt.strftime("%m-%d-%y")
        df.loc[open_idx, "Open Date"] = dates
        df.loc[~open_idx, "Close Date"] = dates

    def set_lvl_names(self, df: pd.DataFrame) -> None:
        open_idx = df[self.ttype_field].isin(self.opening_transactions)
        df.loc[open_idx, "Name2"] = "Opening Transactions"
        df.loc[~open_idx, "Name2"] = "Closing Transactions"

        df.loc[open_idx, "Name3"] = "Purchase of Securities"
        df.loc[~open_idx, "Name3"] = "Sale of Securities"

    def get_date_filtered_transactions(self, df: pd.DataFrame) -> pd.DataFrame:
        df["Date"] = pd.to_datetime(df["Settle Date"])
        start, end = self.report.trxs_start_date, self.end_date
        date_filter = (df.Date <= end) & (df.Date >= start)
        res = df[date_filter].sort_values(by=["Date"])
        if not res.empty:
            return res
        raise ValueError(f"No transactions found for dates: {start} -> {end}!")

    def get_position_long_short_flag(self, df: pd.DataFrame) -> pd.Series:
        """
        Returns a boolean series indicating if a position is long.
        - Logic relies on the `Quantity Balance` and d1g1t transaction quantity field.
        - (qty_bal - trx_qty) >= 0 (long) (qty_bal - trx_qty) < 0 (short)
        - trx_qty needs to be updated for signs since d1g1t qty is always positive regardless of trx_type
        """
        trx_qty = df["Quantity"].copy()
        trx_qty.loc[df[self.ttype_field] == "Sell"] = -df["Quantity"]
        return df["quantity_balance"] - trx_qty >= 0

    def set_rgl_based_on_gain_classification(self, df: pd.DataFrame) -> None:
        """
        Sets the realized-gain/loss to zero based on `gain_classification`.
        If position is long/short and gain classification is not `Capital`
        set rgl to zero.
        """
        posn_is_long = self.get_position_long_short_flag(df)
        gain_clasfn = df["security-udf3"].str.split("|", expand=True)
        long_gain_clasfn, short_gain_clasfn = gain_clasfn[0], gain_clasfn[1]

        long_income_idx = posn_is_long & (long_gain_clasfn != "Capital")
        short_income_idx = ~posn_is_long & (short_gain_clasfn != "Capital")

        zero_gain_idx = long_income_idx | short_income_idx
        for curr_type in ["Local", "Rep. Cur."]:
            df.loc[zero_gain_idx, f"Realized G/L Account {curr_type}"] = 0

    def compute_numerical_fields(self, df: pd.DataFrame) -> None:
        df["Asset Account Local"] = df[self.effect_on_bv_ioc]
        df["Realized G/L Account Local"] = df[self.realized_gl_ioc]

        df["Asset Account Rep. Cur."] = df[self.effect_on_bv_rc]
        df["Realized G/L Account Rep. Cur."] = df[self.realized_gl_rc]
        self.set_rgl_based_on_gain_classification(df)

        for curr_type in ["Local", "Rep. Cur."]:
            df[f"Cash Account {curr_type}"] = -(
                df[f"Asset Account {curr_type}"]
                + df[f"Realized G/L Account {curr_type}"].abs()
            )

        df["Cost Forex"] = df[self.book_value_rc].divide(df[self.book_value_ioc])
        df["Proceeds Forex"] = (
            df["Cash Account Rep. Cur."].divide(df["Asset Account Rep. Cur."])
        ).abs()  # TODO: Needs review.

    def get_capital_trxs_impacting_cash_only(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Get all by,sl transactions (impacting cash only) by
        that end with '-2' and remove all '-1' transactions
        with the same guid as the '-2' transactions.
        """
        df["guid"] = df[self.trx_id].str.rsplit("-", n=1).str[0]
        id2_trx_idx = df[self.trx_id].str.endswith("-2")
        if sum(id2_trx_idx) == 0:  # No id2s exist
            return df
        id2_trxs = df[id2_trx_idx]
        id2_bysl_guids = set(id2_trxs["guid"])
        by_sl_idx = df[self.ttype_field].isin(self.report_ttypes)
        id1_bysl_guids = ~df["guid"].isin(id2_bysl_guids)
        cash_impact_by_sl_idx = by_sl_idx & id1_bysl_guids
        return df[cash_impact_by_sl_idx]

    @staticmethod
    def get_legend(df: pd.DataFrame) -> pd.Series:
        return df["Security Asset Class"].map(lookup.ASSETCLASS_ID_MAP)

    def sort_report(self, df: pd.DataFrame) -> None:
        df.sort_values(by=[self.currency_field, self.ttype_field, "Date"], inplace=True)

    def parse_report(self, df: pd.DataFrame) -> pd.DataFrame:
        for col in self.num_cols:
            df[col] = df[col].astype(float)

        df = self.get_date_filtered_transactions(df)
        self.sort_report(df)
        df[self.currency_field] = df[self.currency_field].map(
            lookup.currency_code_to_currency_name
        )

        self.compute_numerical_fields(df)

        try:
            TaxPrepReport(self.report, report_type="capital").run(df)
        except Exception as err:
            LOG.info(
                f"Got the following error generating the tax prep report 6: {err}..."
            )

        self.set_report_dates(df)
        self.set_lvl_names(df)
        df["legend"] = self.get_legend(df)
        df = df[self.report_columns]
        self.report.recon[self.__class__.__name__] = df.copy()
        res = generate_group_totals(df, self.grp_fields, self.sum_columns)
        return res[self.report_columns]

    def generate_worksheet(self, df) -> None:
        sheet = self.report_sheet

        # Merge cells for the headers
        sheet.merge_cells("A1:O1")
        sheet.merge_cells("A2:O2")
        sheet.merge_cells("A3:O3")

        # Add headers to the merged cells
        sheet["A1"] = "CAPITAL TRANSACTION SUMMARY - SETTLED TRADES"
        sheet["A2"] = self.account_info["name"]
        sheet["A3"] = f"From {self.report.trxs_start_date} to {self.report.end_date}"
        sheet["A4"] = self.report_header

        # Center align the headers
        for row in range(1, 4):
            cell = sheet[f"A{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _first_header():
            cash_real = ["Cash", "Realized"]
            return [""] * 7 + ["Asset"] + cash_real + ["", "Asset", ""] + cash_real

        def _second_header():
            acct_gl = ["Account", "G/L"]
            acct = ["Account"]
            return (
                [""] * 3
                + ["Open", "Close"]
                + [""] * 2
                + acct
                + acct_gl
                + ["Cost"]
                + acct
                + ["Proceeds"]
                + acct_gl
            )

        def _third_header():
            return (
                [""] * 3
                + ["Date"] * 2
                + ["Quantity", "Description"]
                + ["Local"] * 3
                + ["Forex"]
                + ["Rep. Cur."]
                + ["Forex"]
                + ["Rep. Cur."] * 2
            )

        update_cell_range_values(sheet[5:5], _first_header())
        update_cell_range_values(sheet[6:6], _second_header())
        update_cell_range_values(sheet[7:7], _third_header())

        # underline rows A to O
        underline = Border(bottom=Side(style="thin"))
        for cell in sheet[7:7]:  # Rows A7 -> O7
            cell.border = underline

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        # Legend
        last_row = sheet.max_row + 1
        sheet[f"A{last_row}"] = "Legend"
        for i, (key, value) in enumerate(
            lookup.ASSETCLASS_ID_MAP.items(), start=last_row + 1
        ):
            sheet[f"A{i}"] = f"{value} - {key}"

        # Define the font style
        font_style = Font(name="Arial", size=8)

        # Apply the styling to all cells in the sheet
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                cell.font = font_style
            if row_idx >= 8:
                for idx in [5, 7, 8, 11, 13]:
                    row[idx].number_format = "#,##0.00"
                for idx in [10, 12]:
                    row[idx].number_format = "0.0000"
                for idx in [9, 14]:
                    row[idx].number_format = "0.00"
            sheet.row_dimensions[row_idx].height = 12

        # Adjust column widths
        col_widths = {col: 1 for col in "ABCP"}
        col_widths.update({col: 7 for col in "DEFHIJKLMNO"})
        col_widths.update({"G": 33})
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width

    def run(self):
        payload = self.get_calculation_payload()
        try:
            response = self.report.get_calculation("log-details", payload)
            prser = ChartTableFormatter(response, payload)
            parsed_response = prser.parse_data()
            res = self.parse_report(parsed_response)
            self.generate_worksheet(res)
        except Exception as err:
            LOG.warning(
                f"Got the following exception while running Capital Transactions report: {err}..."
            )
