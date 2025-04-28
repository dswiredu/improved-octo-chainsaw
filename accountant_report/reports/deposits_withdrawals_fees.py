import warnings
from pandas.errors import SettingWithCopyWarning

warnings.simplefilter(action="ignore", category=(FutureWarning, SettingWithCopyWarning))

import logging

import pandas as pd
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

from ps.parser import ChartTableFormatter
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport
from claret.accountant_report.report_utils import (
    generate_group_totals,
    set_pagination_size,
    set_trx_payload_custom_date_range,
    get_client_required_date_format,
)
from lookup import currency_code_to_currency_name

LOG = logging.getLogger(__name__)


class DepositsWithdrawalsFees:
    """Class to generate Claret Deposits-Withdrawals-Fees report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.run_dates = [report.start_date, report.end_date]
        self.payload_filename = "deposits_withdrawals_fees_payload.json"
        self.sheet_name = "Deposits-Withdrawals-Fees"
        self.payload_file = report.get_payload_file(self.payload_filename)
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)
        self.report_header = f"Reporting Currency: {self.account_info['reporting_currency_name']}"

        # strings
        self.currency_field = "transaction_currency"
        self.iftn_field = "investor_transaction_type"
        self.comment_field = "comment"
        self.trx_id = "Transaction ID"

        # nums
        self.local = "local"
        self.amount_fx_rate = "amount_fx_rate"
        self.amount_rc = "amount_rc"

        self.grouping_fields = [self.currency_field, self.iftn_field]
        self.num_cols = [self.local, self.amount_rc]

        self.report_columns = [
            self.currency_field,
            self.iftn_field,
            "Date",
            self.comment_field,
            self.local,
            self.amount_fx_rate,
            self.amount_rc,
        ]

    def get_amount_fx_rate(self, df: pd.DataFrame) -> pd.Series:
        return df[self.local].divide(df[self.amount_rc])

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        set_pagination_size(payload)
        set_trx_payload_custom_date_range(
            payload, start=self.report.trxs_start_date, end=self.end_date
        )
        return payload

    def get_date_filtered_transactions(self, df: pd.DataFrame): 
        df["Date"] = pd.to_datetime(df["Settle Date"])
        start, end = self.report.trxs_start_date, self.end_date
        date_filter = (df.Date <= end) & (df.Date >= start)
        res = df[date_filter].sort_values(by=["Date"])
        if not res.empty:
            return res
        raise ValueError(f"No transactions found for dates: {start} -> {end}!")
    
    def split_data_by_currency_conversions(self, df: pd.DataFrame) -> set:
        is_conversion = df[self.iftn_field] == "Currency Conversion"
        line_in_source = df["Transaction ID"].str.endswith("-0")

        conversions = df[is_conversion]
        conversion_guids = set(conversions["GUID"])
        conversion_data = df[df["GUID"].isin(conversion_guids)]
        trxs = df[line_in_source & ~is_conversion]
        return trxs, conversion_data
    
    def get_conversion_legs_together(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        1. Splits dataframe into 1st and 2nd legs for currency conversions
        2. Attaches 2nd leg to 1st leg via guid
        3. Returns resulting dataframe.
        NB: Logic is based on all transactions that end with -0 (1st leg)
        """
        line_in_source = df["Transaction ID"].str.endswith("-0") # 1st leg
        second_leg_idx = df["Transaction ID"].str.endswith("-1") # 2nd leg
        first_leg = df[line_in_source]
        second_leg = df.loc[second_leg_idx, [self.currency_field, "GUID"]]
        second_leg.rename(columns={self.currency_field: "second_leg_currency"}, inplace=True)
        res = first_leg.merge(second_leg, on=["GUID"], validate="1:1") # TODO: Add validation.
        return res
    
    def _get_currency_conversion_pairs(self, df) -> dict:
        currency_pairs = df[[self.currency_field, "second_leg_currency"]]
        unique_pairs = currency_pairs.drop_duplicates(subset=[self.currency_field, "second_leg_currency"])
        return unique_pairs.set_index(self.currency_field)["second_leg_currency"].to_dict()

    def process_currency_conversion_data(self, frame: pd.DataFrame) -> pd.DataFrame:
        """Attaches """
        df = self.get_conversion_legs_together(frame)
        currency_res = []
        currency_pairs = self._get_currency_conversion_pairs(df)
        for first_leg_currency, second_leg_currency in currency_pairs.items():
            currency_df = df[df[self.currency_field] == first_leg_currency]
            first_leg = self._get_currency_conversion_first_leg(currency_df, first_leg_currency)
            second_leg = self._get_currency_conversion_other_leg(first_leg, second_leg_currency)
            currency_legs = pd.concat([first_leg, second_leg], ignore_index=True)
            currency_res.append(currency_legs)
        res = pd.concat(currency_res, ignore_index=True)
        return res

    def _get_currency_conversion_other_leg(
        self, df: pd.DataFrame, currency: str
    ) -> pd.DataFrame:
        res = df.copy()
        local_negative_idx = df[self.local] < 0
        currency_suffix = currency + " (Currency)"

        res[self.amount_rc] = -res[self.amount_rc]
        res[self.local] = res[self.amount_rc]

        res.loc[~local_negative_idx, self.iftn_field] = "Sell of " + currency_suffix
        res.loc[local_negative_idx, self.iftn_field] = "Buy of " + currency_suffix
        res[self.currency_field] = res["second_leg_currency"]
        return res

    def _get_currency_conversion_first_leg(
        self, df: pd.DataFrame, currency: str
    ) -> pd.Series:
        local_negative_idx = df[self.local] < 0
        currency_suffix = currency + " (Currency)"

        res = df[self.iftn_field]
        res.loc[~local_negative_idx] = "Buy of " + currency_suffix
        res.loc[local_negative_idx] = "Sell of " + currency_suffix
        return df

    def parse_report(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        1. Filter transactions between start and end dates
        2. Get grouping totals.
        """
        for col in self.num_cols:
            df[col] = df[col].astype(float)

        df = self.get_date_filtered_transactions(df)
        df["GUID"] = df[self.trx_id].str.split("-").str[0]
        trx, conversions = self.split_data_by_currency_conversions(df)
        if not conversions.empty:
            conversions = self.process_currency_conversion_data(conversions)
        df = pd.concat([trx, conversions], ignore_index=True)
        notnull_idx = df[self.comment_field].notnull() & df[self.iftn_field].notnull()
        df = df[notnull_idx]  # TODO : Ask Pierre
        self.report.recon[self.__class__.__name__] = df.copy()      
        df[self.currency_field] = df[self.currency_field].map(
            currency_code_to_currency_name
        )
        df["Date"] = get_client_required_date_format(df["Date"])
        df[self.amount_fx_rate] = self.get_amount_fx_rate(df)
        res = generate_group_totals(
            frame=df,
            grouping_fields=self.grouping_fields,
            sum_columns=self.num_cols,
            grand_total_sum_columns=[self.amount_rc],
        )
        return res[self.report_columns]

    def generate_worksheet(self, df) -> None:
        sheet = self.report_sheet

        # Merge cells for the headers
        sheet.merge_cells("A1:G1")
        sheet.merge_cells("A2:G2")
        sheet.merge_cells("A3:G3")

        # Add headers to the merged cells
        sheet["A1"] = self.report.translate("DEPOSIT, WITHDRAWALS, AND FEES")
        sheet["A2"] = self.report.translate(self.account_info["name"])
        sheet["A3"] = self.report.translate(f"From {self.report.trxs_start_date} to {self.report.end_date}")
        sheet["A4"] =  self.report.translate(self.report_header)

        # Center align the headers
        for row in range(1, 5):
            cell = sheet[f"A{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _first_header():
            return [""] * 3 + ["Trade", "", "Amount", "Amount"]

        def _second_header():
            return [""] * 2 + ["Date", "Security", "local", "FX Rate", "Rep. Cur."]

        for cell, header in zip(sheet["5:5"], _first_header()):
            cell.value = self.report.translate(header)

        for cell, header in zip(sheet["6:6"], _second_header()):
            cell.value = self.report.translate(header)

        # Underline row G
        underline = Border(bottom=Side(style="thin"))
        for cell in sheet["6:6"]:  # Rows A6 -> G6
            cell.border = underline

        for row in dataframe_to_rows(df, index=False, header=False):
            self.report_sheet.append(row) # TODO: translate row first.
        
        # Define the font style
        font_style = Font(name="Arial", size=8)

        # Apply the styling to all cells in the sheet
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                cell.font = font_style
            if row_idx >= 7:
                for idx in [4,6]: # Cols E,G
                    row[idx].number_format = "#,##0.00"
                row[5].number_format = "0.0000"
            sheet.row_dimensions[row_idx].height = 12
        
        # Adjust column widths
        col_widths = {"A": 1, "B": 1,"D": 33,}
        col_widths.update({col : 7 for col in "CEFG"})
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width



    def run(self):
        payload = self.get_calculation_payload()
        try:
            response = self.report.get_calculation("log-details", payload)
            prser = ChartTableFormatter(response, payload)
            parsed_response = prser.parse_data()
            res = self.parse_report(parsed_response)
            self.generate_worksheet(res)
        except Exception as err:
            LOG.warning(
                f"Got the following exception while running Deposits-Withdrawals-Fees report: {err}..."
            )
