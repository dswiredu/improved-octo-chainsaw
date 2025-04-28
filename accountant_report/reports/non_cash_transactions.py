import logging

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, Border, Side

from ps.parser import ChartTableFormatter
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport

from claret.accountant_report.lookup import (
    currency_code_to_currency_name,
)
from claret.accountant_report.report_utils import (
    get_group_with_no_totals,
    set_pagination_size,
    set_trx_payload_custom_date_range
)

LOG = logging.getLogger(__name__)


class NonCashTransactions:
    """Class to generate Income report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.payload_filename = "non_cash_transactions.json"
        self.sheet_name = "Non-cash Transactions"
        self.payload_file = report.get_payload_file(self.payload_filename)
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)

        # Set the columns to group by
        self.trx_ccy_field = "Transaction Currency"
        self.grp_fields = [self.trx_ccy_field]

        self.num_cols = [
            "d1g1t Transaction Amount (Net)",
            "d1g1t Transaction Amount (Net - Trx Currency)",
            "Value Date Forex",
            "Cost Local",
            "Cost Date Forex",
            "Cost with Forex",
            "d1g1t Transaction Quantity"
        ]
        self.report_columns = [
            "Group",
            "Transaction",
            "Security",
            "Trade date",
            "d1g1t Transaction Quantity",
            "d1g1t Transaction Amount (Net - Trx Currency)",
            "Value Date Forex",
            "d1g1t Transaction Amount (Net)",
            "Original Cost Date",
            "Cost Local",
            "Cost Date Forex",
            "Cost with Forex"
        ]

    @staticmethod
    def get_amount_fx_rate(df: pd.DataFrame) -> pd.Series:
        return df["d1g1t Transaction Amount (Net - Trx Currency)"].divide(
            df["d1g1t Transaction Amount (Net)"]
        )

    def get_report_print_parameters(self):
        year_start, month_start, day_start = self.report.trxs_start_date.split("-")
        report_start = month_start + "-" + day_start + "-" + year_start[-2:]
        year_end, month_end, day_end = self.report.end_date.split("-")
        report_end = month_end + "-" + day_end + "-" + year_end[-2:]
        account_name = self.account_info["name"]
        report_name = "NON-CASH TRANSACTION SUMMARY"
        date_range = "From " + report_start + " To " + report_end
        reporting_currency = (
            "Reporting Currency: "
            + currency_code_to_currency_name[self.account_info["currency"]]
        )
        return account_name, report_name, date_range, reporting_currency

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        set_pagination_size(payload)
        set_trx_payload_custom_date_range(
            payload, start=self.report.trxs_start_date, end=self.end_date
        )
        return payload

    def extract_fx_data(data: dict):
        data =data['results'][0]
        return float(data['close'])

    def get_date_filtered_transactions(
        self, df: pd.DataFrame
    ):  # TODO: Do we need to filter only '-0' transactions?
        df["Date"] = pd.to_datetime(df["Trade date"], format="%m-%d-%y")
        start, end = self.report.trxs_start_date, self.end_date
        date_filter = (df["Date"] <= end) & (df["Date"] >= start)
        res = df[date_filter]
        if not res.empty:
            return res
        raise ValueError(f"No transactions found for dates: {start} -> {end}!")

    def parse_report(self, df: pd.DataFrame) -> pd.DataFrame:
        # TODO: break into several functions
        """
        0. Filter out not book for record and cancelled transactions
        1. Filter transactions between start and end dates
        2. Get grouping totals.
        """
        is_book_of_record = df["Is Book Of Record"].astype(bool)
        is_cancelled = df["Is Cancelled"].astype(bool)
        needs_processing = is_book_of_record & ~is_cancelled
        df = df[needs_processing]
        # Further reduce to just securities with ID starting with pf (for cahs make the id cash as it is coming empty)
        is_id_valied = df["Security ID"].fillna("cash").str.startswith("pf")
        df = df[is_id_valied]
        # TODO: Separate into separate function
        # Now extract only the transfers
        transfer_ins = ["Internal Transfer Security In","Transfer Security In"]
        transfer_outs = ["Internal Transfer Security Out","Transfer Security Out"]
        bool_transfers = df["d1g1t Transaction Type"].isin(transfer_ins + transfer_outs)
        bool_transfer_ins = df["d1g1t Transaction Type"].isin(transfer_ins)
        bool_transfer_outs = df["d1g1t Transaction Type"].isin(transfer_outs)
        df = df[bool_transfers]
        # Now compute the uid that will be needed tonidentify pairs
        df["uid"] = df["Transaction ID"].str[-19:-3]
        # Now add the "Transaction" column
        df["Transaction"] = ""
        df.loc[bool_transfer_outs,"Transaction"] = "Delivery of security"
        df.loc[bool_transfer_ins,"Transaction"] = "Reception of security"
        # now compute the fx rates
        df["Value Date Forex"] = np.nan
        df["Cost Date Forex"] = np.nan
        # need to re-do the bool as we have added more data
        bool_transfer_ins = df["d1g1t Transaction Type"].isin(transfer_ins)
        bool_transfer_outs = df["d1g1t Transaction Type"].isin(transfer_outs)
        df.loc[bool_transfer_outs,"Value Date Forex"] = self.get_amount_fx_rate(df[bool_transfer_outs])
        # Get the FX from the cost date
        bool_date_fx_cost = ~df["Original Cost Date"].isnull()
        # TODO: Expore if there is an API that can retrieve all currencies at ones efficiently
        for item in df[bool_date_fx_cost].index:
            fx_tmp = 1
            sec_ccy = df.loc[item,"Security Currency"]
            acc_ccy = self.account_info["currency"]
            fx_params = {
                "base": sec_ccy,
                "foreign": acc_ccy, 
                "date": df.loc[item,"Original Cost Date"]
            }
            if sec_ccy != acc_ccy:
                fx_tmp = 1/self.report.get_fx_data(fx_params)
            df.loc[item,"Cost Date Forex"] = fx_tmp
        # Now add the cost columnes
        df["Cost Local"] = np.nan
        df["Cost with Forex"] = np.nan
        df.loc[bool_transfer_ins,"Cost Local"] = df.loc[bool_transfer_ins,"d1g1t Transaction Amount (Net - Trx Currency)"]
        df.loc[bool_date_fx_cost,"Cost with Forex"] = df.loc[bool_date_fx_cost,"d1g1t Transaction Amount (Net - Trx Currency)"].divide(df.loc[bool_date_fx_cost,"Cost Date Forex"])
        df.loc[bool_transfer_outs,"Cost with Forex"] = df.loc[bool_transfer_outs,"Effect on Book Value"]
        df.loc[bool_transfer_outs,"Cost Local"] = df.loc[bool_transfer_outs,"Effect on Book Value (TC)"]
        df.loc[bool_transfer_outs,"Cost Date Forex"] = df.loc[bool_transfer_outs,"Cost Local"].divide(df.loc[bool_transfer_outs,"Cost with Forex"])
        # For value date fields we want to copy the amounts of the outs in the ins
        df = df.sort_values(by=["Transaction Currency","Trade date","uid","d1g1t Transaction Type"], ascending=[True,True,True,False])
        df.loc[bool_transfer_outs,"d1g1t Transaction Quantity"] = -1 * np.abs(df.loc[bool_transfer_outs,"d1g1t Transaction Quantity"])
        # Columns to copy
        columns_to_copy = ["d1g1t Transaction Amount (Net - Trx Currency)", "Value Date Forex", "d1g1t Transaction Amount (Net)"]
        # Get first and last index for each "uid" group
        first_indices = df.groupby("uid").head(1).index
        last_indices = df.groupby("uid").tail(1).index
        df.loc[last_indices, columns_to_copy] = np.abs(df.loc[first_indices, columns_to_copy].values)
        # Now extract only the rows wih comments so that they can be added to the dataframe
        bool_comment = df["Comment"] != ""
        df_comment = df[bool_comment].copy()
        # Now set the comment as the security name, and the transaction type chnage it to AA (this is needed to ensure theya re always at the end)
        df_comment["Transaction"] = df_comment["Comment"]
        df_comment["d1g1t Transaction Type"] = "AA"
        # Now concatenate the dfs
        df = pd.concat([df, df_comment])
        # now sort by trade date, then by uid again since we concatenated
        df = df.sort_values(by=["Transaction Currency","Trade date","uid","d1g1t Transaction Type"], ascending=[True,True,True,False])
        # Now that we have concatenated we need to re-index to get back a unique index
        df.index = range(len(df.index))
        for col in self.num_cols:
            df[col] = df[col].astype(float)
        df["Trade date"] = pd.to_datetime(df["Trade date"], format="%Y-%m-%d").dt.strftime("%m-%d-%y")
        df["Original Cost Date"] = pd.to_datetime(df["Original Cost Date"], format="%Y-%m-%d").dt.strftime("%m-%d-%y")
        df = self.get_date_filtered_transactions(df)
        # Now make null the columns for the "comment" row
        comment_bool = ~df["Transaction"].isin(["Delivery of security","Reception of security"])
        df.loc[comment_bool,~df.columns.isin(["Transaction","Transaction Currency"])] = np.nan
        df["Group"] = np.nan
        self.report.recon[self.__class__.__name__] = df.copy()
        res = get_group_with_no_totals(
            df, self.grp_fields, currency_code_to_currency_name
        )
        return res[self.report_columns]

    def save_non_cash_transactions_report(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        account_name: str,
        report_name: str,
        date_range: str,
        reporting_currency: str,
        report_headers: list,
    ) -> None:
        # wb = Workbook()
        ws = self.report_sheet
        # ws.title = sheet_name
        num_header_rows = len(report_headers)
        num_header_cols = len(report_headers[0])
        # Merge the first 4 rows
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=num_header_cols
        )  # Merge cells
        ws.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=num_header_cols
        )  # Merge cells
        ws.merge_cells(
            start_row=3, start_column=1, end_row=3, end_column=num_header_cols
        )  # Merge cells
        ws.merge_cells(
            start_row=4, start_column=1, end_row=4, end_column=num_header_cols
        )  # Merge cells
        # Add centered report name
        # TODO: Can this be done via a loop instead of explicit definition?
        report_name_cell = ws.cell(row=1, column=1)
        report_name_cell.value = report_name
        report_name_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Center alignment
        report_name_cell.font = Font(name="Arial", size=8, bold=True)
        # Add centered account name
        account_name_cell = ws.cell(row=2, column=1)
        account_name_cell.value = account_name
        account_name_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Center alignment
        account_name_cell.font = Font(name="Arial", size=8, bold=True)
        # Add centered date range
        date_range_cell = ws.cell(row=3, column=1)
        date_range_cell.value = date_range
        date_range_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Center alignment
        date_range_cell.font = Font(name="Arial", size=8, bold=True)
        # Add left aligned rporting currency
        reporting_currency_cell = ws.cell(row=4, column=1)
        reporting_currency_cell.value = reporting_currency
        reporting_currency_cell.alignment = Alignment(
            horizontal="left", vertical="center"
        )  # Center alignment
        reporting_currency_cell.font = Font(name="Arial", size=8, bold=True)
        # Add Column Headers (2 rows)
        for row_idx, header_row in enumerate(report_headers, start=5):
            for col_idx, col_value in enumerate(header_row, start=1):
                tmp_cell = ws.cell(row=row_idx, column=col_idx, value=col_value)
                tmp_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )  # Center alignment
                tmp_cell.font = Font(size=8, bold=True)
        # ws.merge_cells(
        #     start_row=5, start_column=7, end_row=5, end_column=9
        # )  # Merge cells
        # ws.merge_cells(
        #     start_row=5, start_column=11, end_row=5, end_column=13
        # )  # Merge cells
        # Define a border style
        thin_bottom_border = Border(bottom=Side(style="thin"))
        # for col in [7, 8, 9, 11, 12, 13]:
        #     cell = ws.cell(row=5, column=col)
        #     cell.border = thin_bottom_border
        for col in range(2, num_header_cols + 1):
            cell = ws.cell(row=7, column=col)
            cell.border = thin_bottom_border
        # Define Custom Styles
        string_style = NamedStyle(
            name="StringStyle",
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="left"),
        )
        english_number_format = "#,##0.00"
        english_fx_number_format = "#,##0.0000"
        number_style = NamedStyle(
            name="NumberStyle",
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="right"),
            number_format=english_number_format,
        )
        fx_number_style = NamedStyle(
            name="NumberStyleFx",
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="right"),
            number_format=english_fx_number_format,
        )
        # Add DataFrame Data with Alignment
        data_start_row = 5 + num_header_rows  # Data starts after the title + headers
        for row_idx, row in enumerate(df.itertuples(index=False), start=data_start_row):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Apply styles based on column
                if (
                    col_idx <= 3
                ):  # | col_idx == num_header_cols:  # first 6 columns are text
                    cell.style = string_style
                else:  # Columns 2-6: Numbers, right-aligned with thousands separator
                    cell.style = number_style
                if col_idx == 7:
                    cell.style = fx_number_style
                if col_idx == 11:
                    cell.style = fx_number_style
                # if col_idx == 9:
                #     cell.style = string_style
        # Set global row height
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 12  # Set height to 12
        # TODO: Make this generic
        columns_to_adjust = {"A": 1, "B": 13, "C": 33, "D": 7, "E": 10, "F": 10,"G":7,"H":10,"I":7,"J":10,"K":7,"L":10}
        for col, width in columns_to_adjust.items():
            ws.column_dimensions[col].width = width
        # wb.save(
        #     "/Users/carlosmartinezamaya/Documents/Work/d1g1t/Code/fe-scripts/claret/accountant_report/out/test.xlsx"
        # )
        # document_name = self.get_report_filename()
        # wb.save(document_name)

    def run(self):
        payload = self.get_calculation_payload()
        try:
            response = self.report.get_calculation("log-details", payload)
            prser = ChartTableFormatter(response, payload)
            parsed_response = prser.parse_data()
            res = self.parse_report(parsed_response)
            (
                account_name,
                report_name,
                date_range,
                reporting_currency,
            ) = self.get_report_print_parameters()
            report_headers = [
                [""]*6 + ["Value"]*2 + [""]*2 + ["Cost"]*2,
                [""]*5 + ["Value"] + ["Date"] + ["With"] + ["Cost"]*2 + ["Date"] + ["With"],
                [
                    "",
                    "Transaction",
                    "Security",
                    "Date",
                    "Quantity",
                    "Local",
                    "Forex",
                    "Forex",
                    "Date",
                    "Local",
                    "Forex",
                    "Forex"
                ]]
            self.save_non_cash_transactions_report(
                res,
                self.sheet_name,
                account_name,
                report_name,
                date_range,
                reporting_currency,
                report_headers
            )
        except Exception as err:
            LOG.error(
                f"Got the following exception while running Non Cash Transactions report: {err}..."
            )
