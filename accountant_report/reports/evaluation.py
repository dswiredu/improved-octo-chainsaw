import logging

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

from ps.parser import ChartTableFormatter
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport
from claret.accountant_report.report_utils import remove_unheld_positions
from claret.accountant_report.report_utils import (
    generate_group_totals,
    update_cell_range_values,
)
from claret.accountant_report.lookup import currency_code_to_currency_name

LOG = logging.getLogger(__name__)


class EvaluationReport:
    """Class to generate Claret evaluation report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.run_dates = [self.report.start_date, self.report.end_date]
        self.payload_filename = "evaluation_payload.json"
        self.payload_file = self.report.get_payload_file(self.payload_filename)
        self.report_headers = f"Reporting Currency: {self.account_info['reporting_currency_name']}"

        self.sheet_name_map = {
            self.report.start_date: "Evaluation - Beginning",
            self.report.end_date: "Evaluation - End",
        }

        # strings
        self.currency_field = "security_currency"
        self.description_field = "Description"
        self.security_type_field = "security_type"
        self.securityid_field = "security_id"
        self.security_asset_class_field = "security_asset_class"
        self.asset_subcategory = "asset_subcategory"

        # Nums
        self.quantity = "quantity"
        self.market_value_settled_dirty_rc = "market_value_settled_dirty_rc"
        self.market_value_settled_clean_rc = "market_value_settled_clean_rc"
        self.market_value_settled_dirty_lc = "market_value_settled_dirty_lc"
        self.market_value_settled_clean_lc = "market_value_settled_clean_lc"
        self.market_value_trade_clean_lc = "market_value_trade_clean_lc"
        self.total_cost_lc = "total_cost_lc"
        self.total_cost_rc = "total_cost_rc"
        self.pct_assets = "pct_assets"
        self.price_lc = "price_lc"

        self.default_asset_class = "Cash & Equivalents"
        self.default_asset_category = "Cash"

        self.num_cols = [
            self.quantity,
            self.market_value_settled_clean_lc,
            self.market_value_settled_clean_rc,
            self.market_value_settled_dirty_lc,
            self.market_value_settled_dirty_rc,
            self.total_cost_lc,
            self.total_cost_rc,
            self.pct_assets,
            self.price_lc,
        ]

        self.grouping_fields = [
            self.currency_field,
            self.security_asset_class_field,
            self.asset_subcategory,
        ]

        self.sum_cols = [
            self.total_cost_lc,
            self.market_value_settled_clean_lc,
            self.total_cost_rc,
            self.market_value_settled_clean_rc,
            self.pct_assets,
        ]

        self.report_columns = [
            self.currency_field,
            self.security_asset_class_field,
            self.asset_subcategory,
            self.quantity,
            self.description_field,
            "unit_cost_lc",
            self.total_cost_lc,
            self.price_lc,
            self.market_value_settled_clean_lc,
            "forex_rate",
            "unit_cost_rc",
            self.total_cost_rc,
            "price_rc",
            self.market_value_settled_clean_rc,
            self.pct_assets,
        ]

        self.rounding_map = {
            "unit_cost_lc": 2,
            "unit_cost_rc": 2,
            self.quantity: 3,
            self.pct_assets: 2,
        }

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        return payload

    def get_unit_cost_columns(self, df: pd.DataFrame) -> None:
        for col in ["rc", "lc"]:
            df[f"unit_cost_{col}"] = df[f"total_cost_{col}"].divide(df[self.quantity])

    def get_accrual_columns(self, df: pd.DataFrame) -> None:
        for col in ["rc", "lc"]:
            df[f"accrual_{col}"] = (
                df[f"market_value_settled_dirty_{col}"]
                - df[f"market_value_settled_clean_{col}"]
            )

    def get_accrued_interest_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        accrual_idx = df[f"accrual_lc"].fillna(0) != 0
        cols = ["accrual_lc", "accrual_rc"]
        accrual = df[accrual_idx]
        accrual = accrual.groupby(self.grouping_fields)[cols].sum().reset_index()
        accrual[self.description_field] = "Accrued Interest"
        rename_cols = {
            "accrual_lc": self.market_value_settled_clean_lc,
            "accrual_rc": self.market_value_settled_clean_rc,
        }
        accrual.rename(columns=rename_cols, inplace=True)
        return pd.concat([df, accrual], ignore_index=True)

    def round_numeric_fields(self, df):
        for field, dp in self.rounding_map.items():
            df[field] = df[field].round(dp)

    def get_forex_rate(self, df: pd.DataFrame) -> pd.Series:
        return df[self.total_cost_lc].divide(df[self.total_cost_rc])

    def set_blank_asset_subcategory(self, df: pd.DataFrame) -> None:
        null_cat = df[self.asset_subcategory].isnull()
        null_class = df[self.security_asset_class_field].isnull()
        currency_idx = df[self.description_field].isin(
            currency_code_to_currency_name.keys()
        )
        df.loc[
            null_cat & currency_idx, self.asset_subcategory
        ] = self.default_asset_category
        df.loc[
            null_class & currency_idx, self.security_asset_class_field
        ] = self.default_asset_class

    def get_missing_report_metrics(self, df: pd.DataFrame) -> None:
        """Retrives metrics not gotten directly from api call: unit costs, forex_rate, price_rc"""
        self.get_unit_cost_columns(df)
        self.get_accrual_columns(df)
        df["forex_rate"] = self.get_forex_rate(df)
        fx_rate = df[self.market_value_settled_clean_lc].divide(
            df[self.market_value_settled_clean_rc]
        )
        df["price_rc"] = df[self.price_lc].divide(fx_rate)
        df[self.pct_assets] = df[self.pct_assets] * 100
    

    def set_incorrect_settle_cash_due_to_accruals(self, df: pd.DataFrame) -> None:
        """Cash is impacted by the Income Accrual number due to opening positions.
        As such we use market_value_trade_clean_lc for just cash.
        """
        cash_idx = df[self.description_field].isin(currency_code_to_currency_name.keys())
        change_cols = [
            self.quantity,
            self.market_value_settled_clean_lc,
            self.market_value_settled_clean_rc,
            self.market_value_settled_dirty_lc,
            self.market_value_settled_dirty_rc,
            self.total_cost_lc,
            self.total_cost_rc,
        ]
        for col in change_cols:
            df.loc[cash_idx, col] = df[self.market_value_trade_clean_lc]

    def parse_report(self, frame: pd.DataFrame) -> pd.DataFrame:
        """Retrives metrics not gotten directly from api call and wrangles data into
        client report format."""
        df = frame[frame["Name"] != "Total"]
        self.set_incorrect_settle_cash_due_to_accruals(df) # TODO: Change once DEVREQ-8154 complete
        df[self.currency_field] = df[self.currency_field].map(
            currency_code_to_currency_name
        )

        for col in self.num_cols:
            df[col] = df[col].astype(float)

        self.get_missing_report_metrics(df)
        df = remove_unheld_positions(df, self.securityid_field)
        self.set_blank_asset_subcategory(df)
        df = self.get_accrued_interest_rows(df)
        self.report.recon[self.sheet_name] = df.copy()
        res = generate_group_totals(df, self.grouping_fields, self.sum_cols)
        self.round_numeric_fields(df)
        return res[self.report_columns]

    def generate_worksheet(self, df, dte) -> None:
        sheet = self.report_sheet

        # Merge cells for the headers
        sheet.merge_cells("A1:O1")
        sheet.merge_cells("A2:O2")
        sheet.merge_cells("A3:O3")

        # Add headers to the merged cells
        sheet["A1"] = "Portfolio Valuation - SETTLED TRADES"
        sheet["A2"] = self.account_info["name"]
        sheet["A3"] = dte
        sheet["A4"] =  self.report_headers

        # Center align the headers
        for row in range(1, 4):
            cell = sheet[f"A{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells("F5:I5")
        sheet.merge_cells("K5:O5")
        for row in ["F5", "K5"]:
            sheet[row].alignment = Alignment(horizontal="center", vertical="center")

        def _first_header():
            return [""] * 5 + ["Local Currency", "", "Reporting Currency"]

        def _second_header():
            metric_list = ["Unit", "Total", "", "Market"]
            return [""] * 5 + metric_list + ["Forex"] + metric_list + ["Pct."]

        def _third_header():
            metric_list = ["Cost"] * 2 + ["Price", "Value"]
            return (
                [""] * 3
                + ["Quantity", "Description"]
                + metric_list
                + ["Rate"]
                + metric_list
                + ["Assets"]
            )

        update_cell_range_values(sheet[5:5], _first_header())
        update_cell_range_values(sheet[6:6], _second_header())
        update_cell_range_values(sheet[7:7], _third_header())

        # underline rows A to O
        underline = Border(bottom=Side(style="thin"))
        for cell in sheet["7:7"]:  # Rows A7 -> O7
            cell.border = underline

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        # Define the font style
        font_style = Font(name="Arial", size=8)

        # Apply the styling to all cells in the sheet
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                cell.font = font_style
            if row_idx >= 8:
                for idx in [3, 6, 8, 11, 13]:
                    row[idx].number_format = "#,##0.000"
                for idx in [5, 7, 9, 10, 12, 14]:
                    row[idx].number_format = "#,##0.00"
                row[4].alignment = Alignment(wrap_text=True)
            sheet.row_dimensions[row_idx].height = 12

        # Adjust column widths
        col_widths = {"E": 30}
        col_widths.update({col: 1 for col in "ABC"})
        col_widths.update({col: 7 for col in "DGILN"})
        col_widths.update({col: 4 for col in "FHJKMO"})
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width

    def run(self):
        payload = self.get_calculation_payload()
        for dte in self.run_dates:
            try:
                LOG.info(f"Running evaluation report for {dte}...")
                payload["settings"]["date"]["date"] = dte
                response = self.report.get_calculation("cph-table", payload)
                prser = ChartTableFormatter(response, payload)
                parsed_response = prser.parse_data()
                self.sheet_name = self.sheet_name_map[dte]
                res = self.parse_report(parsed_response)
                self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)
                self.generate_worksheet(res, dte)
            except Exception as err:
                LOG.warning(
                    f"Got the following exception while running evaluation report for {dte}: {err}..."
                )
