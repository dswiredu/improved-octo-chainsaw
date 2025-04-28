import warnings
from pandas.errors import SettingWithCopyWarning

warnings.simplefilter(action="ignore", category=(FutureWarning, SettingWithCopyWarning))

import logging

import pandas as pd
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

from claret.accountant_report.main import AccountantReport
from claret.accountant_report.report_utils import update_cell_range_values
from lookup import currency_code_to_currency_name

LOG = logging.getLogger(__name__)


class Reconciliation:
    """Class to generate Claret Reconciliation report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.sheet_name = "Reconciliation"
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)
        self.report_header = (
            f"Reporting Currency: {self.account_info['reporting_currency_name']}"
        )

        self.funds_prefix = "Claret Funds - Distributions "

        self.report_columns = [
            "Currency",
            "Name",
            "cash_local_currency",
            "cash_reporting_currency",
            "investment_local_currency",
            "investment_reporting_currency",
            "total_portfolio_local_currency",
            "total_portfolio_reporting_currency",
        ]
        self.string_cols = ["Currency", "Name", "report"]
        self.num_cols = self.report_columns[2:]

        self.parse_map = {
            "Evaluation - Beginning": self.parse_evaluation_begin,
            "IncomeReport": self.parse_income,
            "NonCashTransactions": self.parse_non_cash_trxs,
            "DepositsWithdrawalsFees": self.parse_deposits_withdrwls_fees,
            "CapitalTransactions": self.parse_capital_trxs,
            "Evaluation - End": self.parse_evaluation_end,
        }

    @property
    def reporting_currencies(self):
        currencies = set()
        for _, frame in self.report.recon.items():
            currencies |= set(frame["Currency"])
        return {
            x for x in currencies if x == x
        }  # Todo: There is a nan here. Investigate! cshk2020_hk(HKD)

    def parse_deposits_withdrwls_fees(self, df: pd.DataFrame) -> pd.DataFrame:
        col_map = {
            "investor_transaction_type": "Name",
            "local": "cash_local_currency",
            "amount_rc": "cash_reporting_currency",
        }
        df.rename(columns=col_map, inplace=True)
        df["Currency"] = df["transaction_currency"].map(currency_code_to_currency_name)
        df = df.assign(
            investment_local_currency=0,
            investment_reporting_currency=0,
            total_portfolio_local_currency=df["cash_local_currency"],
            total_portfolio_reporting_currency=df["cash_reporting_currency"],
        )
        return df[self.report_columns]

    @staticmethod
    def get_evaluation_sums(df: pd.DataFrame) -> pd.DataFrame:
        cash_idx = df["asset_subcategory"] == "Cash"
        cash = df[cash_idx]
        cash = cash[["Currency", "cash_local_currency", "cash_reporting_currency"]]

        inv = df[~cash_idx]
        inv.rename(
            columns={
                "cash_local_currency": "investment_local_currency",
                "cash_reporting_currency": "investment_reporting_currency",
            },
            inplace=True,
        )
        inv = inv[
            ["Currency", "investment_local_currency", "investment_reporting_currency"]
        ]
        inv = inv.groupby(["Currency"]).sum().reset_index()
        res = cash.merge(inv, on=["Currency"], how="outer")
        return res.fillna(0)

    def _parse_evaluation_init(self, df: pd.DataFrame) -> pd.DataFrame:
        col_map = {
            "security_currency": "Currency",
            "total_cost_lc": "cash_local_currency",
            "total_cost_rc": "cash_reporting_currency",
        }
        df.rename(columns=col_map, inplace=True)
        df = self.get_evaluation_sums(df)
        df = df.assign(
            total_portfolio_local_currency=df["investment_local_currency"]
            + df["cash_local_currency"],
            total_portfolio_reporting_currency=df["cash_reporting_currency"]
            + df["investment_reporting_currency"],
        )
        return df

    def parse_evaluation_begin(self, df: pd.DataFrame) -> pd.DataFrame:
        res = self._parse_evaluation_init(df)
        res["Name"] = "Balance - Beginning"
        return res[self.report_columns]

    def parse_evaluation_end(self, df: pd.DataFrame) -> pd.DataFrame:
        res = self._parse_evaluation_init(df)
        res["Name"] = "Balance - End"
        return res[self.report_columns]

    def _get_withholding_tax(self, df: pd.DataFrame) -> pd.DataFrame:
        """Returns dataframe that has Withholding tax as an extra set of rows
        to be added to parsed income report.
        """
        taxw = df["Tax Withheld"]
        tax_idx = taxw.notnull() & (taxw != 0)
        if sum(tax_idx) == 0:
            return pd.DataFrame()
        
        tax_col_map = {
            "Name": "Name",
            "Tax Withheld": "cash_reporting_currency",
            "Withholding Tax - Local": "cash_local_currency",
            "Currency": "Currency",
        }

        tax = df.loc[tax_idx, tax_col_map.keys()].copy()
        tax.rename(columns=tax_col_map, inplace=True)
        tax = tax.assign(
            investment_local_currency=0,
            investment_reporting_currency=0,
            total_portfolio_local_currency=tax["cash_local_currency"],
            total_portfolio_reporting_currency=tax["cash_reporting_currency"],
        )
        res = tax.groupby(["Name", "Currency"]).sum().reset_index()
        res[self.num_cols] = -res[self.num_cols]  # From report these are -ve
        res["Name"] = "Withholding Tax - " + res["Name"]
        return res[self.report_columns]

    def parse_income(self, df: pd.DataFrame) -> pd.DataFrame:
        col_map = {
            "d1g1t Transaction Amount (Gross - Trx Currency)": "cash_local_currency",
            "d1g1t Transaction Amount (Gross)": "cash_reporting_currency",
        }
        df.rename(columns=col_map, inplace=True)
        df["Currency"] = df["Transaction Currency"].map(currency_code_to_currency_name)
        df = df.assign(
            investment_local_currency=0,
            investment_reporting_currency=0,
            total_portfolio_local_currency=df["cash_local_currency"],
            total_portfolio_reporting_currency=df["cash_reporting_currency"],
            Name=df["Security Group"] + " (" + df["Country"] + ")",
        )
        tax = self._get_withholding_tax(df)
        df = df[self.report_columns]
        res = pd.concat([df, tax], ignore_index=True)
        return res

    def _get_non_cash_trxs_name_field(self, df: pd.DataFrame) -> pd.DataFrame:
        start_year = self.start_date.split("-")[0]
        end_year = self.end_date.split("-")[0]

        end_year_begin = f"{end_year}-01-01"
        cuttoff_idx = df["Date"] <= end_year_begin
        df.loc[cuttoff_idx, "Name"] = f"{self.funds_prefix}{start_year}"
        df.loc[~cuttoff_idx, "Name"] = f"{self.funds_prefix}{end_year}"
        return df.drop(columns=["Date"])

    def get_non_cash_transaction_summary(self, frame: pd.DataFrame) -> pd.DataFrame:
        for col in self.num_cols:
            frame[col].fillna(0, inplace=True)
        df = frame.groupby(["Currency", "Transaction", "Name"]).sum().reset_index()

        # Filter data for each transaction type
        receipts = df[df["Transaction"] == "Reception of security"]
        delivery = df[df["Transaction"] == "Delivery of security"]

        # Perform subtraction by merging on Currency and Name
        summary = receipts.merge(
            delivery, on=["Currency", "Name"], suffixes=("_rec", "_del"), how="left"
        )

        # Compute the difference for each numeric column
        # Using + for difference cos Delivery has negative numbers
        for col in self.num_cols:
            summary[col] = summary[f"{col}_rec"].fillna(0) + summary[
                f"{col}_del"
            ].fillna(0)

        return summary[self.report_columns]

    def parse_non_cash_trxs(self, df: pd.DataFrame) -> pd.DataFrame:
        non_cash_trxs = ["Delivery of security", "Reception of security"]
        df = df[df["Transaction"].isin(non_cash_trxs)]
        rename_cols = {
            "Settle Date": "Date",
            "Cost Local": "investment_local_currency",
            "Cost with Forex": "investment_reporting_currency",
        }
        df = df[
            [
                "Date",
                "Transaction Currency",
                "Transaction",
                "Cost Local",
                "Cost with Forex",
            ]
        ].rename(columns=rename_cols)
        df["Currency"] = df["Transaction Currency"].map(currency_code_to_currency_name)

        df = df.assign(
            cash_local_currency=0,
            cash_reporting_currency=0,
            total_portfolio_local_currency=df["investment_local_currency"],
            total_portfolio_reporting_currency=df["investment_reporting_currency"],
        )
        res = self._get_non_cash_trxs_name_field(df)
        summary = self.get_non_cash_transaction_summary(res)
        return summary

    def parse_capital_trxs(self, df: pd.DataFrame) -> pd.DataFrame:
        col_map = {
            "transaction_currency": "Currency",
            "Name3": "Name",
            "Cash Account Local": "cash_local_currency",
            "Cash Account Rep. Cur.": "cash_reporting_currency",
            "Asset Account Local": "investment_local_currency",
            "Asset Account Rep. Cur.": "investment_reporting_currency",
        }
        df.rename(columns=col_map, inplace=True)
        df = df.assign(
            total_portfolio_local_currency=df["investment_local_currency"]
            + df["cash_local_currency"],
            total_portfolio_reporting_currency=df["cash_reporting_currency"]
            + df["investment_reporting_currency"],
        )
        return df[self.report_columns]

    def combine_recon_frames(self) -> pd.DataFrame:
        frames = []
        for report, frame in self.report.recon.items():
            func = self.parse_map[report]
            df = func(frame)
            df["report"] = report
            frames.append(df)
        df = pd.concat(frames)

        aggregation = {"report": "first"}
        aggregation.update({x: sum for x in df if x not in self.string_cols})
        res = df.groupby(["Currency", "Name"]).agg(aggregation).reset_index()
        return res

    def get_currency_recon(self, currency: str, frame: pd.DataFrame) -> pd.DataFrame:
        df = frame[frame["Currency"] == currency]

        # Get total row
        total_dict = {"Name": ["Verification"]}
        total_dict.update({x: [pd.NA] for x in ["Currency", "report"]})
        sum_row = pd.DataFrame(total_dict)
        last_col, prior_cols = df.iloc[-1:], df.iloc[:-1]
        sum_row[self.num_cols] = (
            last_col[self.num_cols].fillna(0).sum()
            - prior_cols[self.num_cols].fillna(0).sum()
        )

        df.loc[max(df.index) + 1, :] = pd.NA  # insert empty row at last
        shift_cols = df.columns[1:]
        df[shift_cols] = df[shift_cols].shift(1)

        res = pd.concat([df, sum_row], ignore_index=True)
        res["Currency"] = res["Currency"].fillna(method="ffill")
        res["Currency"] = res["Currency"].fillna("")
        res.loc[res["Currency"] == res["Currency"].shift(1), "Currency"] = ""
        return res

    def parse_recon(self) -> pd.DataFrame:
        df = self.combine_recon_frames()
        df["report"] = pd.Categorical(
            df["report"], ordered=True, categories=self.parse_map.keys()
        )
        df.sort_values(by=["Currency", "report"], inplace=True)
        currency_frames = []
        for currency in sorted(set(df["Currency"])):
            currency_recon = self.get_currency_recon(currency, df)
            currency_frames.append(currency_recon)
        res = pd.concat(currency_frames, ignore_index=True)
        return res[self.report_columns]

    def generate_worksheet(self, df) -> None:
        sheet = self.report_sheet

        # Merge cells for the headers
        sheet.merge_cells("A1:H1")
        sheet.merge_cells("A2:H2")
        sheet.merge_cells("A3:H3")

        # Add headers to the merged cells
        sheet["A1"] = "Period Activity Reconciliation"
        sheet["A2"] = self.account_info["name"]
        sheet["A3"] = f"From {self.report.start_date} to {self.report.end_date}"
        sheet["A6"] = self.report_header

        # underline rows A to H
        underline = Border(bottom=Side(style="thin"))
        for cell in sheet[6:6]:  # Rows
            cell.border = underline

        # Center align the headers
        for row in range(1, 4):
            cell = sheet[f"A{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.merge_cells("C7:D7")
        sheet.merge_cells("E7:F7")
        sheet.merge_cells("G7:H7")

        sheet["C7"] = "Cash"
        sheet["E7"] = "Investments"
        sheet["G7"] = "Total Portfolio"

        # Center align the headers
        for col in ["CEG"]:
            cell = sheet[f"{col}7"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _first_header():
            return [""] * 2 + ["Local", "Reporting"] * 3

        def _second_header():
            return [""] * 2 + ["Currency"] * 6

        update_cell_range_values(sheet[8:8], _first_header())
        update_cell_range_values(sheet[9:9], _second_header())

        for row in dataframe_to_rows(df, index=False, header=False):
            self.report_sheet.append(row)

        # Define the font style
        font_style = Font(name="Arial", size=8)

        # Apply the styling to all cells in the sheet
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                cell.font = font_style
            if row_idx >= 9:
                for idx in range(2, 8):
                    row[idx].number_format = "#,##0.00"
            sheet.row_dimensions[row_idx].height = 12

        # Adjust column widths
        col_widths = {"A": 2, "B": 33}
        col_widths.update({col: 7 for col in "CDEFGH"})
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width

    def run(self):
        try:
            df = self.parse_recon()
            self.generate_worksheet(df)
        except Exception as err:
            LOG.warning(
                f"Got the following exception while running Reconciliation report: {err}..."
            )
