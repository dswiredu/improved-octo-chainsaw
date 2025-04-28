import logging

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

from ps.parser import ChartTableFormatter
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport
from lookup import currency_code_to_currency_name
from claret.accountant_report.report_utils import (
    generate_group_totals,
    set_pagination_size,
    update_cell_range_values,
    set_trx_payload_custom_date_range
)

LOG = logging.getLogger(__name__)
class RealizedGainLossFXReport:
    """Class to generate Realized Gain Loss FX report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.payload_filename = "realized_gain_loss_fx_payload.json"
        self.sheet_name = "Realized Gain-Losses FX"
        self.payload_file = report.get_payload_file(self.payload_filename)
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)
        self.report_header = f"Reporting Currency: {self.account_info['reporting_currency_name']}"

        self.transaction_currency = "transaction_currency"
        self.book_value_rc = "book_value_rc"
        self.trx_amount_rc = "trx_amount_rc"
        self.cash_realized_gl = "cash_realized_gl"

        self.grp_fields = [
            self.transaction_currency,
        ]

        self.num_cols = [
            self.book_value_rc,
            self.trx_amount_rc,
        ]

        self.rename_cols = {
            self.book_value_rc : "Cost",
            self.trx_amount_rc: "Proceeds",
        }

        self.report_columns = [
            self.transaction_currency,
            "Settlement Date",
            "Cost",
            "Proceeds",
            "Gain/Loss",
        ]

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        set_pagination_size(payload)
        set_trx_payload_custom_date_range(
            payload, start=self.report.trxs_start_date, end=self.end_date
        )
        return payload

    def get_date_filtered_transactions(
        self, df: pd.DataFrame
    ):  # TODO: Do we need to filter only '-0' transactions?
        df["Date"] = pd.to_datetime(df["Settle Date"])
        start, end = self.report.trxs_start_date, self.end_date
        date_filter = (df.Date <= end) & (df.Date >= start)
        res = df[date_filter]
        if not res.empty:
            return res
        raise ValueError(f"No transactions found for dates: {start} -> {end}!")

    def parse_report(self, df: pd.DataFrame) -> pd.DataFrame:
        for col in self.num_cols:
            df[col] = df[col].astype(float)

        df = self.get_date_filtered_transactions(df)
        df["Gain/Loss"] = df[self.book_value_rc] - df[self.trx_amount_rc]
        res = generate_group_totals(df, self.grp_fields, self.num_cols)

        res.rename(columns=self.rename_cols, inplace=True)
        res["Settlement Date"] = res["Date"].dt.strftime("%m-%d-%y")
        res[self.transaction_currency] = res[self.transaction_currency].map(
            currency_code_to_currency_name
        )
        return res[self.report_columns]

    def generate_worksheet(self, df) -> None:
        sheet = self.report_sheet

        # Merge cells for the headers
        sheet.merge_cells("A1:E1")
        sheet.merge_cells("A2:E2")
        sheet.merge_cells("A3:E3")

        # Add headers to the merged cells
        sheet["A1"] = "REALIZED GAINS AND LOSSES ON FOREIGN EXCHANGE"
        sheet["A2"] = self.account_info["name"]
        sheet["A3"] = f"From {self.report.trxs_start_date} to {self.report.end_date}"
        sheet["A4"] = self.report_header 

        # Center align the headers
        for row in range(1, 4):
            cell = sheet[f"A{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _first_header():
            return [""] + ["Settlement"] + [""] * 3

        def _second_header():
            return [""] + ["Date", "Cost", "Proceeds", "Gain/Loss"]

        update_cell_range_values(sheet[5:5], _first_header())
        update_cell_range_values(sheet[6:6], _second_header())

        # underline rows A to O
        underline = Border(bottom=Side(style="thin"))
        for cell in sheet[6:6]:  # Rows
            cell.border = underline

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        
        # Define the font style
        font_style = Font(name="Arial", size=8)

        # Apply the styling to all cells in the sheet
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                cell.font = font_style
            if row_idx >= 8:
                for idx in [2, 3, 4]:
                    row[idx].number_format = "#,##0.00"
            sheet.row_dimensions[row_idx].height = 12

        # Adjust column widths
        col_widths = {"A": 7}
        col_widths.update({col: 4 for col in "BCDE"})
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width

    def run(self):
        payload = self.get_calculation_payload()
        try:
            response = self.report.get_calculation("log-details", payload)
            prser = ChartTableFormatter(response, payload)
            parsed_response = prser.parse_data()
            res = self.parse_report(parsed_response)
            self.generate_worksheet(res)
        except Exception as err:
            LOG.warning(
                f"Got the following exception while running Realized Gain Loss report: {err}..."
            )
