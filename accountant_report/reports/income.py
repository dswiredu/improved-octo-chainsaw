import logging

import pandas as pd
from openpyxl.styles import Alignment, Font, NamedStyle, Border, Side

from ps.parser import ChartTableFormatter
from claret.accountant_report.reports.tax_prep import TaxPrepReport
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport

from claret.accountant_report.lookup import (
    currency_code_to_country_name,
    country_name_to_code,
    currency_code_to_currency_name,
)
from claret.accountant_report.report_utils import (
    get_group_totals_with_currency,
    set_pagination_size,
    set_trx_payload_custom_date_range
)

LOG = logging.getLogger(__name__)


class IncomeReport:
    """Class to generate Income report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.payload_filename = "income.json"
        self.sheet_name = "Income"
        self.payload_file = report.get_payload_file(self.payload_filename)
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)

        # Set the columns to group by
        self.country_field = "Country"
        self.trx_ccy_field = "Transaction Currency"
        self.trx_type_field = "Security Group"  # "Transaction Type"
        self.grp_fields = [self.country_field, self.trx_ccy_field, self.trx_type_field]

        self.num_cols = [
            "d1g1t Transaction Amount (Net)",
            "d1g1t Transaction Amount (Net - Trx Currency)",
            "d1g1t Transaction Amount (Gross)",
            "d1g1t Transaction Amount (Gross - Trx Currency)",
            "Tax Withheld",
        ]
        self.sum_cols = [
            "d1g1t Transaction Amount (Net - Trx Currency)",
            "d1g1t Transaction Amount (Gross - Trx Currency)",
            "Withholding Tax - Local",
            "d1g1t Transaction Amount (Net)",
            "d1g1t Transaction Amount (Gross)",
            "Tax Withheld",
        ]
        self.report_columns = [
            self.country_field,
            self.trx_ccy_field,
            self.trx_type_field,
            "Ex-Date",
            "Pay-Date",
            "Description",
            "d1g1t Transaction Amount (Gross - Trx Currency)",
            "Withholding Tax - Local",
            "d1g1t Transaction Amount (Net - Trx Currency)",
            "FX Rate",
            "d1g1t Transaction Amount (Gross)",
            "Tax Withheld",
            "d1g1t Transaction Amount (Net)",
            "Country Code",
        ]

    @staticmethod
    def get_amount_fx_rate(df: pd.DataFrame) -> pd.Series:
        return df["d1g1t Transaction Amount (Net - Trx Currency)"].divide(
            df["d1g1t Transaction Amount (Net)"]
        )

    def get_report_print_parameters(self):
        year_start, month_start, day_start = self.report.trxs_start_date.split("-")
        report_start = month_start + "-" + day_start + "-" + year_start[-2:]
        year_end, month_end, day_end = self.report.end_date.split("-")
        report_end = month_end + "-" + day_end + "-" + year_end[-2:]
        account_name = self.account_info["name"]
        report_name = "INCOME SUMMARY REPORT"
        date_range = "From " + report_start + " To " + report_end
        reporting_currency = (
            "Reporting Currency: "
            + currency_code_to_currency_name[self.account_info["currency"]]
        )
        return account_name, report_name, date_range, reporting_currency

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        set_pagination_size(payload)
        set_trx_payload_custom_date_range(
            payload, start=self.report.trxs_start_date, end=self.end_date
        )
        return payload

    def get_date_filtered_transactions(
        self, df: pd.DataFrame
    ):  # TODO: Do we need to filter only '-0' transactions?
        df["Date"] = pd.to_datetime(df["Pay-Date"], format="%m-%d-%y")
        start, end = self.report.trxs_start_date, self.end_date
        date_filter = (df["Date"] <= end) & (df["Date"] >= start)
        res = df[date_filter]
        if not res.empty:
            return res
        raise ValueError(f"No transactions found for dates: {start} -> {end}!")

    def parse_report(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        0. Filter out not book for record and cancelled transactions
        1. Filter transactions between start and end dates
        2. Get grouping totals.
        """
        is_book_of_record = df["Is Book Of Record"].astype(bool)
        is_cancelled = df["Is Cancelled"].astype(bool)
        needs_processing = is_book_of_record & ~is_cancelled
        df = df[needs_processing]
        # TODO: Separate into separate function
        # Now separate the income transacions from the "-2", which are meant to capture the third level grouping
        # We also need to remove buy and sell that are needed for the income gains
        bool_income = df["d1g1t Transaction Type"].isin(
            ["Instrument Cashflow", "Interest"]
        )
        bool_buy_sell = df["d1g1t Transaction Type"].isin(
            ["Buy", "Sell"]
        )
        bool_for_secondary = ~bool_buy_sell & ~bool_income
        df_secondary = df[bool_for_secondary].copy()
        df_buy_sell = df[bool_buy_sell]
        df = df[bool_income]
        # Now set the index to be the UID needed for searhcing and merging. In this instance is the transaction ID
        # minus the last two characters
        df["uid"] = df["Transaction ID"].str[:-2]
        df_secondary["uid"] = df_secondary["Transaction ID"].str[:-2]
        df_secondary["Security Group"] = df_secondary["Security Name"]
        df_secondary = df_secondary[["uid", "Security Group", "Security ID"]].rename(columns={
            "Security ID": "Security ID2"
        })
        # Now merge them
        df = pd.merge(df, df_secondary, on="uid", how="left")
        # For empty security currency set it to the transaction currency
        bool_sec_ccy = df["Security Currency"].isnull()
        df.loc[bool_sec_ccy,"Security Currency"] = df.loc[bool_sec_ccy,"Transaction Currency"]
        # Apply special logic for cash and equivalents
        bool_id = df["Security ID"].isnull()
        df["cash_id"] = df["Security ID"].str[4:].fillna("cash")
        bool_cash = df["cash_id"].str.startswith("cash")
        bool_asset_subcategory = df["Asset Subcategory"].isin(["Cash","Liquidities"])
        bool_cash_subcategory = bool_id | bool_cash | bool_asset_subcategory
        bool_amount_negative = df["d1g1t Transaction Amount (Net)"] < 0
        bool_amount_positive = df["d1g1t Transaction Amount (Net)"] > 0
        bool_group_income = bool_cash_subcategory & bool_amount_positive
        bool_group_expense = bool_cash_subcategory & bool_amount_negative
        # Now make sure that if there is no group we assign them to the Interest Income (CAD)/(USD)
        df["Security Group"] = df["Security Group"].fillna("Other")
        bool_other_group = df["Security Group"].isin(["Other"])
        df.loc[bool_other_group,"Security Group"] = "Interest Income (" + df.loc[bool_other_group,"Security Currency"] + ")"
        df.loc[bool_group_income,"Security Group"] = "Interest Income (" + df.loc[bool_group_income,"Security Currency"] + ")"
        df.loc[bool_group_expense,"Security Group"] = "Interest Expense (" + df.loc[bool_group_expense,"Security Currency"] + ")"
        # Now set all "pa" transactions that DO NOT have an id2 must show up under the Interest Expense bucket in the Income report.
        bool_pa = df["Transaction Code"].isin(["pa"])
        bool_empty_id2 = df["Security ID2"].isnull()
        bool_pa_and_empty_id2 = bool_pa & bool_empty_id2
        df.loc[bool_pa_and_empty_id2,"Security Group"] = "Interest Expense (" + df.loc[bool_pa_and_empty_id2,"Security Currency"] + ")"
        # Now manipulate the buy/sell to get them in the right places
        df_buy_sell["uid"] = df_buy_sell["Transaction ID"]
        bool_gains = df_buy_sell[["Realized Gain/Loss", "Realized Gains (TC)"]].fillna(0).ne(0).any(axis=1) 
        df_buy_sell = df_buy_sell[bool_gains]
        df_buy_sell[["d1g1t Transaction Amount (Net)","d1g1t Transaction Amount (Net - Trx Currency)"]] = df_buy_sell[["Realized Gain/Loss","Realized Gains (TC)"]]
        df_buy_sell[["d1g1t Transaction Amount (Gross)","d1g1t Transaction Amount (Gross - Trx Currency)"]] = df_buy_sell[["Realized Gain/Loss","Realized Gains (TC)"]]
        # separate buys (short) and sell (long)
        bool_buy  = df_buy_sell["d1g1t Transaction Type"].isin(["Buy"])
        bool_sell  = df_buy_sell["d1g1t Transaction Type"].isin(["Sell"])
        # Splitting the user define 3 column into long/short
        df_buy_sell[["long", "short"]] = df_buy_sell["Security User Defined 3"].str.split("|", expand=True)
        bool_long_income = df_buy_sell["long"].isin(["Income"])
        bool_short_income = df_buy_sell["short"].isin(["Income"])
        bool_long = bool_sell & bool_long_income
        bool_short = bool_buy & bool_short_income
        bool_long_short = bool_long | bool_short
        df_buy_sell["Security Group"] = ""
        df_buy_sell.loc[bool_long_short,"Security Group"] = "Interest Income (" + df_buy_sell.loc[bool_long_short,"Security Currency"] + ")"
        # Now remove the capital ones
        df_buy_sell = df_buy_sell[df_buy_sell["Security Group"] != ""]
        # Now merge them
        df = pd.concat([df, df_buy_sell])
        # For currency we don't get name or country, populate them
        is_currency = df["Security Name"].isnull()
        df.loc[is_currency, "Security Name"] = [
            currency_code_to_currency_name[item]
            for item in df.loc[is_currency, "Transaction Currency"]
        ]
        df.loc[is_currency, "Country"] = [
            currency_code_to_country_name[item]
            for item in df.loc[is_currency, "Transaction Currency"]
        ]
        df["Description"] = df["Security Name"]
        for col in self.num_cols:
            df[col] = df[col].astype(float)
        df["Pay-Date"] = pd.to_datetime(
            df["Settle date"], format="%Y-%m-%d"
        ).dt.strftime("%m-%d-%y")
        df["Ex-Date"] = pd.to_datetime(df["Trade date"], format="%Y-%m-%d").dt.strftime(
            "%m-%d-%y"
        )
        df = self.get_date_filtered_transactions(df)
        
        df["FX Rate"] = self.get_amount_fx_rate(df)
        df["Withholding Tax - Local"] = df["Tax Withheld"].multiply(df["FX Rate"])
        df["Country Code"] = [country_name_to_code[item] for item in df["Country"]]
        self.report.recon[self.__class__.__name__] = df.copy()

        try:
            TaxPrepReport(self.report, "income").run(df)
        except Exception as err:
            LOG.info(
                f"Got the following error generating the tax prep report 3: {err}..."
            )

        res = get_group_totals_with_currency(
            df, self.grp_fields, self.sum_cols, currency_code_to_currency_name
        )
        # res = get_group_totals_with_currency(df, self.grp_fields, self.num_cols)
        #         for col in self.grp_fields:
        # #            row = pd.row([col])
        #             res.loc[res[col].duplicated(), col] = ""

        return res[self.report_columns]

    def save_income_report(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        account_name: str,
        report_name: str,
        date_range: str,
        reporting_currency: str,
        report_headers: list,
    ) -> None:
        # wb = Workbook()
        ws = self.report_sheet
        # ws.title = sheet_name
        num_header_rows = len(report_headers)
        num_header_cols = len(report_headers[0])
        # Merge the first 4 rows
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=num_header_cols
        )  # Merge cells
        ws.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=num_header_cols
        )  # Merge cells
        ws.merge_cells(
            start_row=3, start_column=1, end_row=3, end_column=num_header_cols
        )  # Merge cells
        ws.merge_cells(
            start_row=4, start_column=1, end_row=4, end_column=num_header_cols
        )  # Merge cells
        # Add centered report name
        report_name_cell = ws.cell(row=1, column=1)
        report_name_cell.value = report_name
        report_name_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Center alignment
        report_name_cell.font = Font(name="Arial", size=8, bold=True)
        # Add centered account name
        account_name_cell = ws.cell(row=2, column=1)
        account_name_cell.value = account_name
        account_name_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Center alignment
        account_name_cell.font = Font(name="Arial", size=8, bold=True)
        # Add centered date range
        date_range_cell = ws.cell(row=3, column=1)
        date_range_cell.value = date_range
        date_range_cell.alignment = Alignment(
            horizontal="center", vertical="center"
        )  # Center alignment
        date_range_cell.font = Font(name="Arial", size=8, bold=True)
        # Add left aligned rporting currency
        reporting_currency_cell = ws.cell(row=4, column=1)
        reporting_currency_cell.value = reporting_currency
        reporting_currency_cell.alignment = Alignment(
            horizontal="left", vertical="center"
        )  # Center alignment
        reporting_currency_cell.font = Font(name="Arial", size=8, bold=True)
        # Add Column Headers (2 rows)
        for row_idx, header_row in enumerate(report_headers, start=5):
            for col_idx, col_value in enumerate(header_row, start=1):
                tmp_cell = ws.cell(row=row_idx, column=col_idx, value=col_value)
                tmp_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )  # Center alignment
                tmp_cell.font = Font(size=8, bold=True)
        ws.merge_cells(
            start_row=5, start_column=7, end_row=5, end_column=9
        )  # Merge cells
        ws.merge_cells(
            start_row=5, start_column=11, end_row=5, end_column=13
        )  # Merge cells
        # Define a border style
        thin_bottom_border = Border(bottom=Side(style="thin"))
        for col in [7, 8, 9, 11, 12, 13]:
            cell = ws.cell(row=5, column=col)
            cell.border = thin_bottom_border
        for col in range(4, num_header_cols + 1):
            cell = ws.cell(row=6, column=col)
            cell.border = thin_bottom_border
        # Define Custom Styles
        string_style = NamedStyle(
            name="StringStyle",
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="left"),
        )
        english_number_format = "#,##0.00"
        english_fx_number_format = "#,##0.0000"
        number_style = NamedStyle(
            name="NumberStyle",
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="right"),
            number_format=english_number_format,
        )
        fx_number_style = NamedStyle(
            name="NumberStyleFx",
            font=Font(name="Arial", size=8),
            alignment=Alignment(horizontal="right"),
            number_format=english_fx_number_format,
        )
        # Add DataFrame Data with Alignment
        data_start_row = 5 + num_header_rows  # Data starts after the title + headers
        for row_idx, row in enumerate(df.itertuples(index=False), start=data_start_row):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Apply styles based on column
                if (
                    col_idx <= 6
                ):  # | col_idx == num_header_cols:  # first 6 columns are text
                    cell.style = string_style
                else:  # Columns 2-6: Numbers, right-aligned with thousands separator
                    cell.style = number_style
                if col_idx == 10:
                    cell.style = fx_number_style
                if col_idx == num_header_cols:
                    cell.style = string_style
        # Set global row height
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 12  # Set height to 12
        # TODO: Make this generic
        columns_to_adjust = {"A": 1, "B": 1, "C": 1, "D": 7, "E": 7, "F": 33}
        for col, width in columns_to_adjust.items():
            ws.column_dimensions[col].width = width
        # wb.save(
        #     "/Users/carlosmartinezamaya/Documents/Work/d1g1t/Code/fe-scripts/claret/accountant_report/out/test.xlsx"
        # )
        # document_name = self.get_report_filename()
        # wb.save(document_name)

    def run(self):
        payload = self.get_calculation_payload()
        try:
            response = self.report.get_calculation("log-details", payload)
            prser = ChartTableFormatter(response, payload)
            parsed_response = prser.parse_data()
            res = self.parse_report(parsed_response)
            (
                account_name,
                report_name,
                date_range,
                reporting_currency,
            ) = self.get_report_print_parameters()
            report_headers = [
                [
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Local Currency",
                    "",
                    "",
                    "",
                    "Reporting Currency",
                    "",
                    "",
                    "",
                ],
                [
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Gross",
                    "Withholding",
                    "Net",
                    "",
                    "Gross",
                    "Withholding",
                    "Net",
                    "",
                ],
                [
                    "",
                    "",
                    "",
                    "Ex-Date",
                    "Pay-Date",
                    "Description",
                    "Amount",
                    "Tax",
                    "Amount",
                    "FX Rate",
                    "Amount",
                    "Tax",
                    "Amount",
                    "Country",
                ],
            ]
            self.save_income_report(
                res,
                self.sheet_name,
                account_name,
                report_name,
                date_range,
                reporting_currency,
                report_headers,
            )
        except Exception as err:
            LOG.error(
                f"Got the following exception while running Income report: {err}..."
            )
