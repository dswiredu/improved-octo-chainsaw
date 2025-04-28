import logging

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

from ps.parser import ChartTableFormatter
from utils.main_utils import get_json
from claret.accountant_report.main import AccountantReport
from claret.accountant_report.report_utils import (
    generate_group_totals,
    set_pagination_size,
    update_cell_range_values,
    set_trx_payload_custom_date_range
)

LOG = logging.getLogger(__name__)


class RealizedGainLossReport:
    """Class to generate Realized Gain Loss report."""

    def __init__(self, report: AccountantReport):
        self.report = report
        self.account_info = report.account_info
        self.output_location = report.output_location
        self.start_date = report.start_date
        self.end_date = report.end_date
        self.payload_filename = "realized_gain_loss_payload.json"
        self.sheet_name = "Realized Gain-Losses"
        self.payload_file = report.get_payload_file(self.payload_filename)
        self.report_sheet = self.report._workbook.create_sheet(self.sheet_name)
        self.report_header = (
            f"Reporting Currency: {self.account_info['reporting_currency_name']}"
        )

        self.security_asset_class_field = "security_asset_class"
        self.trx_currency = "transaction_currency"
        self.security_currency = "security_currency"
        self.bv_effect_ioc = "effect_on_book_value_ioc"
        self.bv_effect_rc = "effect_on_book_value_rc"
        self.bv_rc = "book_value_rc"
        self.realized_gl_rc = "realized_gl_rc"
        self.quantity = "Quantity"
        self.trx_amount_trx_ccy = "trx_amount_trx_ccy"
        self.trx_amount_ioc = "trx_amount_ioc"
        self.trx_amount_rc = "trx_amount_rc"

        self.grp_fields = [
            self.security_asset_class_field,
        ]

        self.num_cols = [
            self.realized_gl_rc,
            self.bv_effect_ioc,
            self.bv_effect_rc,
            self.bv_rc,
            self.quantity,
            self.trx_amount_trx_ccy,
            self.trx_amount_rc,
        ]

        self.sum_cols = [
            "Cost",
            "Proceeds",
            "Price Gain",
            "FX Gain",
            "Gain/Loss",
        ]

        self.rename_cols = {
            "Date": "Settlement Date",
            "Security Name": "Description",
            self.bv_rc: "Cost",
            self.realized_gl_rc: "Gain/Loss",
            self.trx_amount_rc: "Proceeds",
        }

        self.report_columns = [
            self.security_asset_class_field,
            "Settlement Date",
            "Quantity",
            "Description",
            "Cost",
            "Proceeds",
            "Price Gain",
            "FX Gain",
            "Gain/Loss",
            "Country",
        ]

    def get_calculation_payload(self) -> dict:
        """Get calculation payload and update calculation currency."""
        payload = get_json(path=self.payload_file)
        payload["settings"]["currency"] = self.account_info["currency"]
        payload["control"]["selected_entities"]["accounts_or_positions"] = [
            [self.account_info["entity_id"]]
        ]
        set_pagination_size(payload)
        set_trx_payload_custom_date_range(
            payload, start=self.report.trxs_start_date, end=self.end_date
        )
        return payload

    def get_date_filtered_transactions(
        self, df: pd.DataFrame
    ):  # TODO: Do we need to filter only '-0' transactions?
        df["Date"] = pd.to_datetime(df["Settle Date"])
        start, end = self.report.trxs_start_date, self.end_date
        date_filter = (df.Date <= end) & (df.Date >= start)
        res = df[date_filter]
        if not res.empty:
            return res
        raise ValueError(f"No transactions found for dates: {start} -> {end}!")

    def get_country(self, df: pd.DataFrame) -> pd.Series:
        currency = df[self.trx_currency]
        first_two = currency.str[:2]
        return first_two.str.lower()

    def get_trx_fx_rates(self, df: pd.DataFrame):
        fx_col_map = {
            "trade_date": "date",
            "security_currency": "base",
            "transaction_currency": "foreign",
        }
        fx_cols = fx_col_map.keys()
        fx_rows = df[fx_cols].drop_duplicates(subset=fx_cols).copy()
        fx_rows.rename(columns=fx_col_map, inplace=True)
        fx_input = fx_rows[fx_rows["foreign"] != fx_rows["base"]].reset_index()
        if not fx_input.empty:
            fx_input["fx_rate"] = pd.NA
            for idx, row in fx_input.iterrows():
                row["fx_rate"] = 1 / self.report.get_fx_data(
                    row
                )  # Todo: Come back. is this always division?
                fx_input.iloc[idx] = row
            trx_col_map = {v: k for k, v in fx_col_map.items()}
            fx_input.rename(columns=trx_col_map, inplace=True)
        return fx_input.drop(columns=["index"])

    def get_trx_amount_ioc(self, df: pd.DataFrame):
        """
        Convert trx_amount_trx_ccy to instr-ccy using
        retrieved fx-rates for all transactions on trade-dates
        """
        return df[self.trx_amount_trx_ccy] * df["fx_rate"]

    def compute_price_gain(self, df: pd.DataFrame) -> pd.Series:
        implied_bv_fx = df[self.bv_effect_rc].divide(df[self.bv_effect_ioc])
        asset_appreciation_amt = df[self.trx_amount_ioc] - df[self.bv_effect_ioc]
        return asset_appreciation_amt * implied_bv_fx

    def get_trxs_with_fx_rates(self, df: pd.DataFrame) -> pd.DataFrame:
        """Retrieves all transaction fx rates and attaches it to trxs data."""
        fx = self.get_trx_fx_rates(df)
        if not fx.empty:
            merge_cols = ["trade_date", "security_currency", "transaction_currency"]
            df = df.merge(fx, on=merge_cols, how="left")
            df["fx_rate"].fillna(1, inplace=True)
        else:
            df["fx_rate"] = 1
        return df

    def parse_report(self, df: pd.DataFrame) -> pd.DataFrame:
        for col in self.num_cols:
            df[col] = df[col].astype(float)

        df = self.get_date_filtered_transactions(df)
        df = self.get_trxs_with_fx_rates(df)
        df[self.trx_amount_ioc] = self.get_trx_amount_ioc(df)
        df["Price Gain"] = self.compute_price_gain(df)
        df.rename(columns=self.rename_cols, inplace=True)
        df["FX Gain"] = df["Price Gain"] - df["Gain/Loss"]
        df["Settlement Date"] = df["Settlement Date"].dt.strftime("%m-%d-%y")

        res = generate_group_totals(df, self.grp_fields, self.sum_cols)
        res["Country"] = self.get_country(res)
        return res[self.report_columns]

    def generate_worksheet(self, df) -> None:
        sheet = self.report_sheet

        # Merge cells for the headers
        sheet.merge_cells("A1:J1")
        sheet.merge_cells("A2:J2")
        sheet.merge_cells("A3:J3")

        # Add headers to the merged cells
        sheet["A1"] = "REALIZED GAINS AND LOSSES - SETTLED TRADES"
        sheet["A2"] = self.account_info["name"]
        sheet["A3"] = f"From {self.report.start_date} to {self.report.end_date}"
        sheet["A4"] = self.report_header

        # Center align the headers
        for row in range(1, 4):
            cell = sheet[f"A{row}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _first_header():
            return [""] + ["Settlement"] + [""] * 8

        def _second_header():
            return [""] + [
                "Date",
                "Quantity",
                "Description",
                "Cost",
                "Proceeds",
                "Price Gain",
                "FX Gain",
                "Gain/Loss",
                "Country",
            ]

        update_cell_range_values(sheet[5:5], _first_header())
        update_cell_range_values(sheet[6:6], _second_header())

        # underline rows A to O
        underline = Border(bottom=Side(style="thin"))
        for cell in sheet[6:6]:  # Rows
            cell.border = underline

        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)

        # Define the font style
        font_style = Font(name="Arial", size=8)

        # Apply the styling to all cells in the sheet
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                cell.font = font_style
            if row_idx >= 8:
                for idx in [4, 8]:
                    row[idx].number_format = "#,##0.00"
                row[2].number_format = "0.0000"
            sheet.row_dimensions[row_idx].height = 12

        # Adjust column widths
        col_widths = {"A": 1, "D": 33, "J": 4}
        col_widths.update({col: 7 for col in "BCEFGHI"})
        for col, width in col_widths.items():
            sheet.column_dimensions[col].width = width

    def run(self):
        payload = self.get_calculation_payload()
        try:
            response = self.report.get_calculation("log-details", payload)
            prser = ChartTableFormatter(response, payload)
            parsed_response = prser.parse_data()
            res = self.parse_report(parsed_response)
            self.generate_worksheet(res)
        except Exception as err:
            LOG.warning(
                f"Got the following exception while running Realized Gain Loss report: {err}..."
            )
